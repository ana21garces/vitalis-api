"""
Reportes descargables del panel de administración (Excel y PDF).

Cada reporte se arma en dos pasos independientes:

1. Una función `construir_*` consulta la base y devuelve una `Tabla`: título,
   subtítulo, columnas y filas ya convertidas a texto. Ahí vive toda la lógica.
2. Los renderers `render_excel` / `render_pdf` toman esa `Tabla` y producen el
   archivo. No saben nada del contenido, así que sirven para cualquier reporte.

Los índices y niveles salen de las columnas ya calculadas en la base (las mismas
que ve el usuario), no se recalculan aquí.
"""
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO, StringIO

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.encuesta_hplp import EncuestaHplp
from app.models.user import User, UserRole

# Cuentas del sistema: no responden la encuesta. Se excluyen de los reportes que
# miden participación o progresión (sí aparecen en el reporte de usuarios).
ROLES_PROFESIONALES = [
    UserRole.ADMIN.value,
    UserRole.CAPELLAN.value,
    UserRole.ACTIVIDAD_FISICA.value,
    UserRole.RESPONSABILIDAD_SALUD.value,
    UserRole.RELACIONES_INTERPERSONALES.value,
    UserRole.MANEJO_ESTRES.value,
    UserRole.NUTRICION.value,
]

ROLE_LABELS = {
    "student": "Usuario",
    "admin": "Administrador",
    "capellan": "Psicología Positiva",
    "actividad_fisica": "Actividad física",
    "responsabilidad_salud": "Responsabilidad en salud",
    "relaciones_interpersonales": "Relaciones interpersonales",
    "manejo_estres": "Manejo del estrés",
    "nutricion": "Nutrición",
}

# (prefijo de columna en BD, etiqueta legible). El orden es el del reporte.
DIMENSIONES = [
    ("ri", "Relaciones interpersonales"),
    ("n", "Nutrición"),
    ("rs", "Responsabilidad en salud"),
    ("af", "Actividad física"),
    ("me", "Manejo del estrés"),
    ("pp", "Psicología positiva"),
]
DIM_POR_CLAVE = {
    "relaciones_interpersonales": "ri",
    "nutricion": "n",
    "responsabilidad_salud": "rs",
    "actividad_fisica": "af",
    "manejo_estres": "me",
    "psicologia_positiva": "pp",
}

NIVELES = ["Pobre", "Moderado", "Bueno", "Excelente"]

TIPOS_VALIDOS = {"usuarios", "participacion", "progresion", "distribucion"}
FORMATOS_VALIDOS = {"excel", "pdf", "csv"}


@dataclass
class Tabla:
    titulo: str
    subtitulo: str
    columnas: list[str]
    filas: list[list[str]]
    # Subconjunto de columnas (por índice) que se muestra en el PDF cuando la
    # tabla es demasiado ancha para imprimirse. None = todas.
    columnas_pdf: list[int] | None = None
    nota_pdf: str | None = None


# ── Helpers de consulta ───────────────────────────────────────────────────────

def _mapa_por_usuario(db: Session, extremo) -> dict:
    """Una encuesta por usuario según `extremo` (func.max/func.min de id).

    max -> la más reciente; min -> la primera (la línea base de esa persona).
    """
    sub = (
        db.query(EncuestaHplp.usuario_id, extremo(EncuestaHplp.id).label("eid"))
        .group_by(EncuestaHplp.usuario_id)
        .subquery()
    )
    filas = db.query(EncuestaHplp).join(sub, EncuestaHplp.id == sub.c.eid).all()
    return {e.usuario_id: e for e in filas}


def _conteo_por_usuario(db: Session) -> dict:
    filas = (
        db.query(EncuestaHplp.usuario_id, func.count(EncuestaHplp.id))
        .group_by(EncuestaHplp.usuario_id)
        .all()
    )
    return {uid: n for uid, n in filas}


def _fecha(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d") if dt else ""


def _fecha_hora(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else "Sin registro"


def _indice_nivel(enc: EncuestaHplp | None, clave: str) -> tuple[str, str]:
    """(índice, nivel) de una encuesta para el ámbito `clave` ('global' o prefijo)."""
    if enc is None:
        return "", ""
    if clave == "global":
        ind, niv = enc.indice_global, enc.nivel_global
    else:
        ind, niv = getattr(enc, f"{clave}_indice"), getattr(enc, f"{clave}_nivel")
    return ("" if ind is None else str(ind)), (niv or "")


def _tendencia(nivel_base: str, nivel_actual: str) -> str:
    if nivel_base not in NIVELES or nivel_actual not in NIVELES:
        return "—"
    b, a = NIVELES.index(nivel_base), NIVELES.index(nivel_actual)
    return "Subió" if a > b else "Bajó" if a < b else "Igual"


# ── Reporte 1: usuarios de la plataforma ──────────────────────────────────────

def construir_usuarios(db: Session, rol: str = "todos") -> Tabla:
    q = db.query(User)
    if rol == "usuarios":
        q = q.filter(User.role.notin_(ROLES_PROFESIONALES))
    elif rol == "profesionales":
        q = q.filter(User.role.in_(ROLES_PROFESIONALES))
    usuarios = q.order_by(User.full_name).all()

    ultimas = _mapa_por_usuario(db, func.max)

    # «Origen» distingue las cuentas creadas por el administrador (is_verified,
    # nacen verificadas) de las que se registraron solas por el formulario. No
    # hay verificación de correo, así que ese es el único significado real.
    base = ["Nombre", "Email", "Facultad", "Programa", "Tipo", "Sexo", "Rol",
            "Origen", "Activo", "Fecha de registro", "Hizo encuesta",
            "Última encuesta", "Índice global", "Nivel global"]
    cols_dim = []
    for _, etiqueta in DIMENSIONES:
        cols_dim += [f"{etiqueta} (índice)", f"{etiqueta} (nivel)"]
    columnas = base + cols_dim

    filas = []
    for u in usuarios:
        enc = ultimas.get(u.id)
        ind_g, niv_g = _indice_nivel(enc, "global")
        fila = [
            u.full_name, u.email, u.facultad or "", u.program or "",
            u.tipo_usuario or "", u.sexo or "", ROLE_LABELS.get(u.role, u.role),
            "Admin" if u.is_verified else "Registro", "Sí" if u.is_active else "No",
            _fecha(u.created_at), "Sí" if enc else "No",
            _fecha(enc.fecha_respuesta) if enc else "", ind_g, niv_g,
        ]
        for prefijo, _ in DIMENSIONES:
            ind, niv = _indice_nivel(enc, prefijo)
            fila += [ind, niv]
        filas.append(fila)

    # El PDF muestra el resumen; el detalle por dimensión queda en el Excel.
    cols_pdf = [0, 1, 2, 3, 4, 5, 6, 7, 10, 13]  # hasta "Nivel global"
    etiqueta_rol = {"todos": "todos los roles", "usuarios": "solo usuarios",
                    "profesionales": "solo profesionales"}[rol]
    return Tabla(
        titulo="Usuarios de la plataforma",
        subtitulo=f"{len(filas)} usuarios · {etiqueta_rol} · generado el {_fecha(datetime.now())}",
        columnas=columnas,
        filas=filas,
        columnas_pdf=cols_pdf,
        nota_pdf="El detalle por dimensión está disponible en la versión Excel.",
    )


# ── Reporte 2: participación en las encuestas ─────────────────────────────────

SEGMENTOS = {
    "todas": "todos los usuarios",
    "hizo_base": "completaron la encuesta inicial",
    "con_seguimiento": "completaron al menos un seguimiento",
    "ambas": "completaron inicial y seguimiento",
    "solo_una": "completaron solo una encuesta",
    "ninguna": "no han respondido",
}


def construir_participacion(db: Session, segmento: str = "todas") -> Tabla:
    usuarios = (
        db.query(User)
        .filter(User.role.notin_(ROLES_PROFESIONALES))
        .order_by(User.full_name)
        .all()
    )
    conteo = _conteo_por_usuario(db)
    ultimas = _mapa_por_usuario(db, func.max)

    def pasa(n: int) -> bool:
        if segmento == "todas":
            return True
        if segmento == "hizo_base":
            return n >= 1
        if segmento == "con_seguimiento":
            return n >= 2
        if segmento == "ambas":
            return n >= 2
        if segmento == "solo_una":
            return n == 1
        if segmento == "ninguna":
            return n == 0
        return True

    columnas = ["Nombre", "Email", "Facultad", "Programa", "Tipo", "Sexo",
                "Encuestas completadas", "Hizo inicial", "Seguimientos",
                "Última encuesta", "Consentimiento aceptado",
                "Índice global actual", "Nivel global actual"]

    filas = []
    for u in usuarios:
        n = conteo.get(u.id, 0)
        if not pasa(n):
            continue
        enc = ultimas.get(u.id)
        ind_g, niv_g = _indice_nivel(enc, "global")
        seguimientos = max(0, n - 1)
        filas.append([
            u.full_name, u.email, u.facultad or "", u.program or "",
            u.tipo_usuario or "", u.sexo or "", str(n), "Sí" if n >= 1 else "No",
            str(seguimientos), _fecha(enc.fecha_respuesta) if enc else "",
            _fecha_hora(enc.consentimiento_aceptado_en) if enc else "",
            ind_g, niv_g,
        ])

    return Tabla(
        titulo="Participación en las encuestas",
        subtitulo=f"{len(filas)} usuarios · {SEGMENTOS.get(segmento, '')} · generado el {_fecha(datetime.now())}",
        columnas=columnas,
        filas=filas,
    )


# ── Reporte 3: progresión de niveles ──────────────────────────────────────────

def _ambitos(dimension: str) -> list[tuple[str, str]]:
    """Ámbitos (clave, etiqueta) a incluir según la dimensión pedida."""
    if dimension == "global":
        return [("global", "Global")]
    if dimension == "todas":
        return [("global", "Global")] + [(p, e) for p, e in DIMENSIONES]
    prefijo = DIM_POR_CLAVE[dimension]
    etiqueta = dict((p, e) for p, e in DIMENSIONES)[prefijo]
    return [(prefijo, etiqueta)]


def construir_progresion(db: Session, dimension: str = "global", nivel: str | None = None) -> Tabla:
    usuarios = (
        db.query(User)
        .filter(User.role.notin_(ROLES_PROFESIONALES))
        .order_by(User.full_name)
        .all()
    )
    bases = _mapa_por_usuario(db, func.min)
    ultimas = _mapa_por_usuario(db, func.max)

    ambitos = _ambitos(dimension)
    # El nivel se filtra por el ámbito principal: la dimensión elegida, o el
    # global cuando se piden todas o el global.
    clave_filtro = ambitos[-1][0] if dimension not in ("global", "todas") else "global"

    base = ["Nombre", "Email", "Facultad", "Programa", "Sexo"]
    cols_dim = []
    for _, etiqueta in ambitos:
        cols_dim += [f"{etiqueta}: nivel inicial", f"{etiqueta}: índice inicial",
                     f"{etiqueta}: nivel actual", f"{etiqueta}: índice actual",
                     f"{etiqueta}: tendencia"]
    columnas = base + cols_dim

    filas = []
    for u in usuarios:
        base_enc = bases.get(u.id)
        ult_enc = ultimas.get(u.id)
        if base_enc is None:  # solo quienes tienen al menos una encuesta
            continue
        if nivel:
            _, niv_actual_filtro = _indice_nivel(ult_enc, clave_filtro)
            if niv_actual_filtro != nivel:
                continue
        fila = [u.full_name, u.email, u.facultad or "", u.program or "", u.sexo or ""]
        for clave, _ in ambitos:
            ind_b, niv_b = _indice_nivel(base_enc, clave)
            ind_a, niv_a = _indice_nivel(ult_enc, clave)
            fila += [niv_b, ind_b, niv_a, ind_a, _tendencia(niv_b, niv_a)]
        filas.append(fila)

    # En PDF, cuando se piden todas las dimensiones, se muestra solo el ámbito
    # global para que la tabla quepa; el detalle completo va en el Excel.
    cols_pdf = None
    nota = None
    if dimension == "todas":
        cols_pdf = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]  # base + primer ámbito (Global)
        nota = "El detalle por cada dimensión está disponible en la versión Excel."

    etiqueta_dim = "todas las dimensiones" if dimension == "todas" else (
        "índice global" if dimension == "global" else ambitos[-1][1])
    filtro_nivel = f" · nivel actual: {nivel}" if nivel else ""
    return Tabla(
        titulo="Progresión de niveles",
        subtitulo=f"{len(filas)} usuarios · {etiqueta_dim}{filtro_nivel} · generado el {_fecha(datetime.now())}",
        columnas=columnas,
        filas=filas,
        columnas_pdf=cols_pdf,
        nota_pdf=nota,
    )


# ── Reporte 4: distribución por niveles ───────────────────────────────────────

def construir_distribucion(db: Session, dimension: str = "global") -> Tabla:
    encuestables = {
        u.id for u in db.query(User.id).filter(User.role.notin_(ROLES_PROFESIONALES)).all()
    }
    # Cada usuario cuenta una vez, con su encuesta más reciente (su estado actual).
    ultimas = {
        uid: enc for uid, enc in _mapa_por_usuario(db, func.max).items()
        if uid in encuestables
    }
    ambitos = _ambitos(dimension)

    def conteos(clave: str) -> dict[str, int]:
        c = {n: 0 for n in NIVELES}
        for enc in ultimas.values():
            _, niv = _indice_nivel(enc, clave)
            if niv in c:
                c[niv] += 1
        return c

    total_con_encuesta = len(ultimas)

    if len(ambitos) == 1:
        clave, etiqueta = ambitos[0]
        c = conteos(clave)
        columnas = ["Nivel", "Usuarios", "Porcentaje"]
        filas = [
            [n, str(c[n]),
             f"{round(c[n] / total_con_encuesta * 100, 1)}%" if total_con_encuesta else "0%"]
            for n in NIVELES
        ]
        titulo_dim = "índice global" if clave == "global" else etiqueta
    else:
        columnas = ["Dimensión"] + NIVELES + ["Con encuesta"]
        filas = []
        for clave, etiqueta in ambitos:
            c = conteos(clave)
            filas.append([etiqueta] + [str(c[n]) for n in NIVELES] + [str(total_con_encuesta)])
        titulo_dim = "todas las dimensiones"

    return Tabla(
        titulo="Distribución por niveles",
        subtitulo=f"{total_con_encuesta} usuarios con encuesta · {titulo_dim} · generado el {_fecha(datetime.now())}",
        columnas=columnas,
        filas=filas,
    )


# ── Renderers ─────────────────────────────────────────────────────────────────

VERDE = "16A34A"


def render_csv(tabla: Tabla) -> bytes:
    """CSV plano con una fila por registro, para analizar en SPSS, R o Excel.

    A diferencia del Excel, no lleva título ni formato: solo el encabezado y los
    datos, que es lo que esperan los programas de estadística. Se escribe con
    BOM (`utf-8-sig`) porque Excel en Windows, sin él, rompe las tildes.
    """
    import csv

    salida = StringIO()
    escritor = csv.writer(salida, delimiter=";", lineterminator="\r\n")
    escritor.writerow(tabla.columnas)
    escritor.writerows(tabla.filas)
    return salida.getvalue().encode("utf-8-sig")


def render_excel(tabla: Tabla) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ncols = max(1, len(tabla.columnas))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=tabla.titulo).font = Font(bold=True, size=14, color="1F2937")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=tabla.subtitulo).font = Font(size=10, color="6B7280")

    encabezado_fila = 4
    relleno = PatternFill("solid", fgColor=VERDE)
    for j, nombre in enumerate(tabla.columnas, start=1):
        c = ws.cell(row=encabezado_fila, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = relleno
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, fila in enumerate(tabla.filas, start=encabezado_fila + 1):
        for j, valor in enumerate(fila, start=1):
            ws.cell(row=i, column=j, value=valor)

    # Ancho aproximado por el contenido más largo de cada columna.
    for j in range(1, ncols + 1):
        largo = len(str(tabla.columnas[j - 1]))
        for fila in tabla.filas:
            if j - 1 < len(fila):
                largo = max(largo, len(str(fila[j - 1])))
        ws.column_dimensions[get_column_letter(j)].width = min(45, max(10, largo + 2))

    ws.freeze_panes = ws.cell(row=encabezado_fila + 1, column=1)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def render_pdf(tabla: Tabla) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # Proyecta a las columnas del PDF si la tabla es demasiado ancha.
    if tabla.columnas_pdf is not None:
        idx = tabla.columnas_pdf
        columnas = [tabla.columnas[i] for i in idx]
        filas = [[f[i] for i in idx] for f in tabla.filas]
    else:
        columnas, filas = tabla.columnas, tabla.filas

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("t", parent=estilos["Title"], fontSize=16,
                                   textColor=colors.HexColor("#1F2937"), spaceAfter=2)
    estilo_sub = ParagraphStyle("s", parent=estilos["Normal"], fontSize=9,
                                textColor=colors.HexColor("#6B7280"), spaceAfter=8)
    estilo_celda = ParagraphStyle("c", parent=estilos["Normal"], fontSize=7, leading=8)
    estilo_head = ParagraphStyle("h", parent=estilos["Normal"], fontSize=7,
                                 leading=8, textColor=colors.white, fontName="Helvetica-Bold")
    estilo_nota = ParagraphStyle("n", parent=estilos["Normal"], fontSize=7.5,
                                 textColor=colors.HexColor("#94A3B8"), spaceBefore=8)

    datos = [[Paragraph(str(c), estilo_head) for c in columnas]]
    for fila in filas:
        datos.append([Paragraph(str(v), estilo_celda) for v in fila])

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=landscape(A4),
        leftMargin=1 * cm, rightMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm,
        title=tabla.titulo,
    )
    ancho_util = doc.width
    col_width = ancho_util / max(1, len(columnas))

    tabla_pdf = Table(datos, colWidths=[col_width] * len(columnas), repeatRows=1)
    tabla_pdf.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{VERDE}")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elementos = [
        Paragraph(tabla.titulo, estilo_titulo),
        Paragraph(tabla.subtitulo, estilo_sub),
        tabla_pdf,
    ]
    if not filas:
        elementos.append(Spacer(1, 6))
        elementos.append(Paragraph("No hay datos para los filtros seleccionados.", estilo_nota))
    if tabla.nota_pdf:
        elementos.append(Paragraph(tabla.nota_pdf, estilo_nota))

    doc.build(elementos)
    return bio.getvalue()


# ── Orquestación ──────────────────────────────────────────────────────────────

def generar(db: Session, tipo: str, *, rol: str, segmento: str,
            dimension: str, nivel: str | None) -> Tabla:
    if tipo == "usuarios":
        return construir_usuarios(db, rol)
    if tipo == "participacion":
        return construir_participacion(db, segmento)
    if tipo == "progresion":
        return construir_progresion(db, dimension, nivel)
    if tipo == "distribucion":
        return construir_distribucion(db, dimension)
    raise ValueError(f"Tipo de reporte desconocido: {tipo}")


def render(tabla: Tabla, formato: str) -> tuple[bytes, str, str]:
    """Devuelve (contenido, media_type, extensión) para el formato pedido."""
    if formato == "csv":
        return render_csv(tabla), "text/csv; charset=utf-8", "csv"
    if formato == "excel":
        return (
            render_excel(tabla),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    return render_pdf(tabla), "application/pdf", "pdf"
